# coding:utf-8

import time
import os
import copy
import logging
from openpyxl import load_workbook

#-------------------------------------------------------------------------------
#-------------------------------------------------------------------------------
# 类名称：Screenshot
# 类的目的：截取浏览器界面图片
# 假设：无
# 影响：无
# 输入：使用时需要传driver对象
# 返回值：无
# 创建者：lingy
# 创建时间：2018/7/13
# 修改者：
# 修改原因：
# 修改时间:
#-------------------------------------------------------------------------------
def save_screen_picture(driver):
    picture_dir = os.path.dirname(os.getcwd())+r'\output\screenshot\\'#获取存放图片的文件夹路径
    picture_file = time.strftime('%Y-%m-%d_%H%M%S',time.localtime(time.time()))
    picture_path = picture_dir + picture_file +'.png'
    log = InsertLog()
    try:
        driver.get_screenshot_as_file(picture_path)
        time.sleep(1)
        log.info(u'截图成功，存放路径"%s"'%(picture_path,))
    except Exception as msg:
        log.error(u'截图失败：%s'%msg)


#-------------------------------------------------------------------------------
#-------------------------------------------------------------------------------
# 类名称：InsertLog
# 类的目的：写日志
# 假设：无
# 影响：无
# 输入：无
# 返回值：无
# 创建者：lingy
# 创建时间：2017/11/18
# 修改者：
# 修改原因：
# 修改时间:
#-------------------------------------------------------------------------------
#当前文件路径
cur_path = os.path.dirname(os.path.realpath(__file__))
# log_path是存放日志的路径
log_path = os.path.join(os.path.dirname(cur_path), r'output\\logs')
# 如果不存在这个logs文件夹，就自动创建一个
if not os.path.exists(log_path):os.mkdir(log_path)

class InsertLog():
    def __init__(self):
        # 文件的命名
        self.logname = os.path.join(log_path, '%s.log'%time.strftime('%Y_%m_%d'))
        self.logger = logging.getLogger()
        self.logger.setLevel(logging.DEBUG)
        # 日志输出格式
        self.formatter = logging.Formatter('[%(asctime)s - %(funcName)s line: %(lineno)3d] - %(levelname)s: %(message)s')

    def __console(self, level, message):
        # 创建一个FileHandler，用于写到本地
        fh = logging.FileHandler(self.logname, 'a')  # 追加模式  这个是python2的
        # fh = logging.FileHandler(self.logname, 'a', encoding='utf-8')  # 这个是python3的
        fh.setLevel(logging.DEBUG)
        fh.setFormatter(self.formatter)
        self.logger.addHandler(fh)

        # 创建一个StreamHandler,用于输出到控制台
        ch = logging.StreamHandler()
        ch.setLevel(logging.DEBUG)
        ch.setFormatter(self.formatter)
        self.logger.addHandler(ch)

        if level == 'info':
            self.logger.info(message)
        elif level == 'debug':
            self.logger.debug(message)
        elif level == 'warning':
            self.logger.warning(message)
        elif level == 'error':
            self.logger.error(message)
        # 这两行代码是为了避免日志输出重复问题
        self.logger.removeHandler(ch)
        self.logger.removeHandler(fh)
        # 关闭打开的文件
        fh.close()

    def debug(self, message):
        self.__console('debug', message)

    def info(self, message):
        self.__console('info', message)

    def warning(self, message):
        self.__console('warning', message)

    def error(self, message):
        self.__console('error', message)

#-------------------------------------------------------------------------------
#-------------------------------------------------------------------------------
# 函数/过程名称：GetSkipScripts
# 函数/过程的目的：获取不需要执行的模块名字
# 假设：无
# 影响：无
# 输入：无
# 返回值：无
# 创建者：lingy
# 创建时间：2018/06/02
# 修改者：
# 修改原因：
# 修改时间:
#-------------------------------------------------------------------------------

def GetSkipScripts(FilePath):
    try:
        m = []
        wb = load_workbook(FilePath)
        ws = wb['ScriptPath']
        rowcount = ws.max_row
        for i in range(1,rowcount+1):
            cellvalue = ws.cell(row=i,column=3).value
            if cellvalue=='False':
                modulename = ws.cell(row=i,column=2).value
                m.append(modulename)
        wb.close()
        return m
    except BaseException as msg:
        log = InsertLog()
        log.error(msg)

#-------------------------------------------------------------------------------
#-------------------------------------------------------------------------------
# 函数/过程名称：GetSkipScripts
# 函数/过程的目的：获取不需要执行的模块名字
# 假设：无
# 影响：无
# 输入：无
# 返回值：无
# 创建者：lingy
# 创建时间：2018/07/17
# 修改者：
# 修改原因：
# 修改时间:
#-------------------------------------------------------------------------------

def GetSkipTestCases(FilePath):
    try:
        t = []
        wb = load_workbook(FilePath)
        sheels = wb.sheetnames
        #print sheels
        for i in sheels:
            ws = wb[i]
            rowcount = ws.max_row
            for j in range(1,rowcount+1):
                cellvalue = ws.cell(row=j,column=7).value
                if cellvalue=='False':
                    testcasename = ws.cell(row=j,column=1).value
                    t.append(testcasename)
        wb.close()
        return t
    except BaseException as msg:
        log = InsertLog()
        log.error(msg)

#-------------------------------------------------------------------------------
#-------------------------------------------------------------------------------
# 函数/过程名称：get_test_suite
# 函数/过程的目的：筛选出并去除不需要执行的脚本和用例
# 假设：无
# 影响：无
# 输入：discover加载所有测试脚本；m需要去除的脚本；t需要去除的用例
# 返回值：需要执行的测试用例集
# 创建者：lingy
# 创建时间：2018/07/17
# 修改者：
# 修改原因：
# 修改时间:
#-------------------------------------------------------------------------------
def get_test_suite(discover,m,t):
    #筛选出并去除不需要执行的脚本
    suite_m = copy.deepcopy(discover)
    for i in range(len(m)):
        for j in range(discover._tests.__len__()):
            d = discover._tests[j]
            if m[i] in str(d):
                suite_m._tests.remove(d)
    #筛选出并去除不需要执行的用例
    suite_c = copy.deepcopy(suite_m)
    for i in range(len(t)):
        for j in range(suite_m._tests.__len__()):
            s_m =  suite_m._tests[j]
            for z in range(s_m._tests.__len__()):
                s_c = s_m._tests[z]
                for k in range(s_c._tests.__len__()):
                    s_t = s_c._tests[k]
                    if t[i] == s_t._testMethodName:
                        suite_c._tests[j]._tests[z]._tests.remove(s_t)
    return suite_c



if __name__ == '__main__':
    log = InsertLog()
    log.info("1111")
